import os
import openpyxl
import csv
import pandas as pd
import numpy as np
from melt_maps import *

method_shortname_dict = {2022: method_shortname_rmb_2022,
                         2024: method_shortname_rmb_2023}
rl_values_dict = {2022: rl_values_rmb_2022,
                  2024: rl_values_rmb_2023}
mdl_values_dict = {2022: mdl_values_rmb_2022,
                   2024: mdl_values_rmb_2023}
parameter_unit_shortname_dict = {2022: parameter_unit_shortname_rmb_2022,
                                 2024: parameter_unit_shortname_rmb_2023}
parameter_unit_symbol_dict = {2022: parameter_unit_symbol_rmb_2022,
                              2024: parameter_unit_symbol_rmb_2023}
parameter_type_dict = {2022: parameter_type_rmb_2022,
                       2024: parameter_type_rmb_2023}

def process_xlsx_files(input_folder):
    for file_name in os.listdir(input_folder):
        if file_name.endswith(".xlsx"):
            input_file_path = os.path.join(input_folder, file_name)
            output_file_path = os.path.join(input_folder, os.path.splitext(file_name)[0] + ".csv")
            process_workbook(input_file_path, output_file_path)

def process_workbook(input_file_path, output_file_path):
    wb = openpyxl.load_workbook(input_file_path)
    combined_data = combine_tabs(wb)
    with open(output_file_path, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        for row in combined_data:
            writer.writerow(row)

def combine_tabs(wb):
    combined_data = []
    original_headers = ['LAKENAME', 'DNRID', 'Date', 'SITE', 'Depth-m', 'STRAT-M', 'EPIL-Y,N', 
                        'DO-mg/L', 'T-°C', 'SEC-m', 'Chl-a', 'TP (MG/L)', 'Ortho-P (mg/L)', 'PH-SU', 
                        'TURB-NTU', 'CL-MG/L', 'SPCON-UMHOS', 'NO3N-MG/L', 'NH3N-MG/L', 
                        'TKN-MG/L', 'TPM-MG/L', 'OPM-MG/L']
    seen_header = False
    
    for sheet_name in wb.sheetnames:
        # Skip certain sheet names
        if sheet_name not in ["Metadata", "ALL", "Como", "ALLDATA", "CRWD"]:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                # If header hasn't been added, add it explicitly
                if not seen_header:
                    combined_data.append(original_headers)  # Use the specified headers
                    seen_header = True
                # Add rows that match specified lake names
                if row[0] in ["COMO", "CROSBY", "Little Crosby", "LOEB", "MCCARRON"]:
                    combined_data.append(row)
    return combined_data

def process_csv(input_folder, output_folder):
    try:
        filecount = 0
        for filename in os.listdir(input_folder):
            
        
            if filename.endswith(".csv"):
                # Read CSV file and create pandas dataframe
                input_filepath = os.path.join(input_folder, filename)
                df = pd.read_csv(input_filepath, encoding='latin-1')
                year = int(df['Date'][0][:4])
                
                # Apply melt function to convert wide data to long data
                post_rows = melt_dataframe(df, filename, output_folder, year)

                filecount += 1
                print(f"Successfully processed {filename}")
                print(f"Total rows before processing: {df.shape[0]}")
                print(f"Total rows after processing: {post_rows}")
                os.remove(input_filepath)
        print(f"All processed data saved to: {os.path.abspath(output_folder)}")
        print(f"Total files processed: {filecount}")

    except Exception as e:
        print(f"Error occurred during processing: {e}")

def melt_dataframe(df, filename, output_folder, year):
    melted_dfs = []
    #Modify the columns and process the data for years greater than 2021
    try:
        def rename_depth_columns(df):
            for column in df.columns:
                if 'depth' in column.lower():
                    df.rename(columns={column: 'Depth (m)'}, inplace=True)
                if 'strat' in column.lower():
                    df.rename(columns={column: 'Strat (m)'}, inplace=True)

        # Call the function to rename depth-related columns
        rename_depth_columns(df)
        site_index = df.columns.get_loc("SITE")
        # Extract the column headers to the right of the "SITE" column
        param_columns = df.columns[site_index + 1:].tolist()
        other_columns = other_columns_rmb
        
        df['Sample depth'] = (df['Depth (m)'] *  3.281).round(2)
        # Initialize a counter for sample number
        sample_number = 0

        # Iterate over the rows of the DataFrame
        for index, row in df.iterrows():
            # Check if the current depth is 0
            if row['Depth (m)'] == 0 or row['Depth (m)'] == 0.5 or row['Depth (m)'] == 0.6:
                # Reset the sample number to 0
                sample_number = 1
            else:
                # Increment the sample number by 1
                sample_number += 1
            
            # Assign the current sample number to the 'Sample number' column
            df.at[index, 'Sample number'] = sample_number

        for param in param_columns:
            melted_df = pd.melt(df, id_vars=other_columns, value_vars=[param], var_name="Parameter type", value_name="Parameter value")
            melted_dfs.append(melted_df)
        
        # Concatenate all melted DataFrames into a single DataFrame
        combined_df = pd.concat(melted_dfs)    
        
        combined_df['Date'] = pd.to_datetime(combined_df['Date'], format='%Y-%m-%d %H:%M:%S', errors='coerce')
        
        combined_df['Measuring program shortname'] = 'Lake'
        combined_df.rename(columns={'SITE':'Sampling location', 'Date':'DateTime'}, inplace=True)
        
        # format DateTime column
        combined_df['DateTime'] = pd.to_datetime(combined_df['DateTime'])
        combined_df['DateTime'] = combined_df['DateTime'].dt.strftime('%m/%d/%Y')
        split_date = combined_df['DateTime'].str.split('/')
        reformatted_date = split_date.str[2] + split_date.str[0].str.zfill(2) + split_date.str[1].str.zfill(2)
        
        # Add and format all necessary columns required for import
        combined_df['Reformatted Date'] = reformatted_date
        combined_df['Reformatted LAKENAME'] = combined_df['LAKENAME'].map(reformatted_lakename_rmb)
        combined_df['Sampling number'] = combined_df['Reformatted Date'].astype(str) + combined_df['Reformatted LAKENAME'].astype(str) + combined_df['Sampling location'].astype(str)
        combined_df.drop(columns=['Reformatted Date', 'Reformatted LAKENAME'], inplace=True)
        combined_df['Station number'] = combined_df['LAKENAME'].astype(str) + combined_df['Sampling location'].astype(str)
        combined_df['Station number'] = np.where(combined_df['Station number'].str.contains('Little'), 'CRWD50', combined_df['Station number'])
        combined_df['Station number'].replace('MCCARRON101', 'MCCARRONS101', inplace=True)
        combined_df['Parameter value'].replace({'Y': '1','N': '0'}, inplace=True)         
        combined_df['Sample id'] = combined_df['Sample number']
        combined_df['Sample datetime'] = combined_df['DateTime']
        combined_df['Sample replicate flag'] = '0'
        combined_df['Parameter method'] = ''
        combined_df['Parameter method short name'] = combined_df['Parameter type'].map(method_shortname_dict[year])
        combined_df['Parameter unit shortname'] = combined_df['Parameter type'].map(parameter_unit_shortname_dict[year])
        combined_df['Parameter unit symbol'] = combined_df['Parameter type'].map(parameter_unit_symbol_dict[year])
        combined_df['Parameter status'] = ''
        combined_df['MDL'] = combined_df['Parameter type'].map(mdl_values_dict[year])
        combined_df['RL'] = combined_df['Parameter type'].map(rl_values_dict[year])
        combined_df['Parameter type'] = combined_df['Parameter type'].map(parameter_type_dict[year])
        combined_df.drop(columns=['LAKENAME'], inplace=True)
        combined_df = combined_df[column_order]
        combined_df.dropna(subset=['Parameter value'], inplace=True)
        combined_df_no_sample_numbers = combined_df.copy()
        combined_df_no_sample_numbers['Sample number'] = ''
        
        # Save formatted csv
        filename = os.path.splitext(filename)[0]
        
        # Output with sample numbers
        output_name =  f'{output_folder}{filename}_processed.csv'
        combined_df.to_csv(output_name, index=False)

        # Output without sample numbers
        output_name =  f'{output_folder}{filename}_no_sample_numbers_processed.csv'
        combined_df_no_sample_numbers.to_csv(output_name, index=False)
        
    except Exception as e:
        print(f"Error occurred during processing for year {year}: {e}")
    return combined_df.shape[0]

input_folder = "input/"
output_folder = "output/"

# Call the function to process CSV files and save the combined data
process_xlsx_files(input_folder)
process_csv(input_folder, output_folder)
